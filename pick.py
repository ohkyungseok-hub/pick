# make_picking_sheet.py
# 사용법:
#   python make_picking_sheet.py "원본.xlsx" "결과.xlsx"
#
# 요구 라이브러리:
#   pip install pandas openpyxl

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter


def excel_col_to_zero_index(col_letter: str) -> int:
    """Excel column letter (e.g., 'A', 'J') -> pandas zero-based index"""
    col_letter = col_letter.strip().upper()
    n = 0
    for ch in col_letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {col_letter}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def build_picking_sheet(
    src_path: str,
    out_path: str,
    # 원본 파일 컬럼 매핑(요청하신 최종 확정 버전)
    colmap=None,
):
    if colmap is None:
        # 원본 기준:
        # 상품연동코드 J, 주문상품 K, 옵션 L,
        # 주문수량 N, 주문회원 Q, 주소 V, 주문요청사항 W
        colmap = {
            "상품연동코드": "J",
            "주문상품": "K",
            "옵션": "L",
            "주문수량": "N",
            "주문회원": "Q",
            "주소": "V",
            "주문요청사항": "W",
        }

    # 1) 원본 읽고 필요한 열만 추출
    df = pd.read_excel(src_path)

    idxs = [excel_col_to_zero_index(colmap[k]) for k in [
        "상품연동코드", "주문상품", "옵션", "주문수량", "주문회원", "주소", "주문요청사항"
    ]]
    df_sel = df.iloc[:, idxs].copy()
    df_sel.columns = ["상품연동코드", "주문상품", "옵션", "주문수량", "주문회원", "주소", "주문요청사항"]

    # 2) 정렬: 주소(오름), 상품연동코드(내림)
    df_sorted = df_sel.sort_values(
        by=["주소", "상품연동코드"],
        ascending=[True, False],
        kind="mergesort"
    )

    # 3) 주소별 합계행 추가
    #    합계행: 주문상품="합계", 주문수량=주소별 sum, 주소=해당 주소
    out_chunks = []
    for addr, g in df_sorted.groupby("주소", sort=False):
        out_chunks.append(g)

        subtotal = {c: "" for c in df_sorted.columns}
        subtotal["주문상품"] = "합계"
        # 주문수량이 숫자가 아닐 수 있어 안전하게 숫자 변환
        subtotal_qty = pd.to_numeric(g["주문수량"], errors="coerce").fillna(0).sum()
        subtotal["주문수량"] = subtotal_qty
        subtotal["주소"] = addr

        out_chunks.append(pd.DataFrame([subtotal]))

    df_final = pd.concat(out_chunks, ignore_index=True)

    # 4) 엑셀 저장
    df_final.to_excel(out_path, index=False)

    # 5) 인쇄/서식 설정(openpyxl)
    wb = load_workbook(out_path)
    ws = wb.active

    # 헤더: 굵게 + 줄바꿈
    header_font = Font(bold=True)
    header_align = Alignment(wrap_text=True, vertical="center")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.alignment = header_align

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    addr_col = headers["주소"]

    # 긴 텍스트 줄바꿈 + 위쪽 정렬
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        for name in ["주문상품", "옵션", "주소", "주문요청사항"]:
            ws.cell(r, headers[name]).alignment = wrap_top

    # 열 너비(필요하면 여기 값만 조정)
    widths = {
        "상품연동코드": 18,
        "주문상품": 60,
        "옵션": 50,
        "주문수량": 10,
        "주문회원": 18,
        "주소": 50,
        "주문요청사항": 40,
    }
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(headers[name])].width = w

    # 인쇄 설정: 가로 + 1페이지 너비 맞춤 + 제목행 반복
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    # 주소 바뀔 때마다 페이지 나누기
    ws.row_breaks.brk = []
    if ws.max_row >= 2:
        prev_addr = ws.cell(2, addr_col).value
        for r in range(3, ws.max_row + 1):
            curr_addr = ws.cell(r, addr_col).value
            if curr_addr != prev_addr:
                ws.row_breaks.append(Break(id=r - 1))  # 이전 행 뒤에서 끊기
                prev_addr = curr_addr

    # 인쇄영역 지정
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(out_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print('사용법: python make_picking_sheet.py "원본.xlsx" "결과.xlsx"')
        sys.exit(1)

    src = sys.argv[1]
    out = sys.argv[2]
    build_picking_sheet(src, out)
    print(f"완료: {out}")
